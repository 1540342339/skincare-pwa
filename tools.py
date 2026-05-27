# tools.py
import re
import os
import json
import ast
import textwrap
import time
import threading
import shutil
import logging
import traceback
from datetime import datetime as dt

import psutil
import operator
import pyperclip
import requests
from PIL import Image
import qrcode
from deep_translator import GoogleTranslator
from langchain.tools import tool
from langchain_core.tools import Tool
from knowledge_base import load_knowledge_base

# ----- 日志配置 -----
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    filename=os.path.join(os.path.dirname(os.path.abspath(__file__)), "app.log"),
    encoding="utf-8",
    filemode="a",
)
logger = logging.getLogger("小想")

TAVILY_API_KEY = os.getenv("TAVILY_API_KEY")

if not TAVILY_API_KEY:
    logger.warning("未设置 TAVILY_API_KEY，联网搜索将使用 DuckDuckGo 作为备选")

# ----- 路径沙箱 -----
PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))

def _validate_path(filepath: str) -> str:
    """校验并返回绝对路径，拒绝项目目录之外的文件操作。返回规范化的绝对路径。"""
    abs_path = os.path.abspath(filepath)
    if not os.path.isabs(filepath):
        abs_path = os.path.normpath(os.path.join(PROJECT_ROOT, filepath))
    else:
        abs_path = os.path.normpath(abs_path)
    if not abs_path.startswith(PROJECT_ROOT + os.sep) and abs_path != PROJECT_ROOT:
        allowed_dirs = [
            os.path.join(PROJECT_ROOT, "output"),
            os.path.join(PROJECT_ROOT, "backups"),
        ]
        if not any(abs_path.startswith(d + os.sep) or abs_path == d for d in allowed_dirs):
            raise PermissionError(f"不允许访问项目目录外的路径: {filepath}")
    return abs_path

OUTPUT_DIR = "output"

def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def auto_backup(filepath: str) -> str:
    """在修改前自动备份文件，返回备份路径"""
    if not os.path.exists(filepath):
        return ""
    backup_dir = "backups"
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = dt.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"{os.path.basename(filepath)}.{timestamp}.bak")
    shutil.copy2(filepath, backup_path)
    return backup_path

# ----- 工具1：联网搜索 -----
@tool
def web_search(query: str) -> str:
    """联网实时搜索最新信息。适合查询新闻、实时动态、网络文章、网页内容等需要从互联网获取的信息。查询新闻、实时资讯请用此工具。"""
    if TAVILY_API_KEY:
        try:
            from langchain_tavily import TavilySearch
            search = TavilySearch(tavily_api_key=TAVILY_API_KEY, max_results=5, search_depth="advanced",
                                  include_answer=True, include_raw_content=False, include_images=False,
                                  days=7)
            raw = search.invoke(query)
            if isinstance(raw, dict):
                answer = raw.get('answer', '')
                if answer:
                    return f"📌 {answer}"
                sources = raw.get('results', [])
                if sources:
                    return _format_search_results(sources, "Tavily")
            answer = getattr(raw, 'answer', '')
            if answer:
                return f"📌 {answer}"
            sources = getattr(raw, 'results', None) or []
            if sources:
                return _format_search_results(sources, "Tavily")
        except Exception as e:
            logger.warning(f"Tavily 搜索失败: {e}，将降级到 DuckDuckGo")
    try:
        from duckduckgo_search import DDGS
        with DDGS() as ddgs:
            raw = list(ddgs.text(query, max_results=5))
        if raw:
            return "\n".join([f"- {r.get('body', '')[:200]}" for r in raw if r.get('body')])
        return "未找到相关信息。"
    except Exception as e:
        logger.error(f"DuckDuckGo 搜索失败: {e}")
        return f"搜索失败：{str(e)}"

def _format_search_results(sources: list, engine: str) -> str:
    """统一格式化搜索结果，每条标注来源。"""
    formatted = []
    for i, s in enumerate(sources[:5]):
        if isinstance(s, dict):
            title = (s.get('title', '') or '')[:80]
            content = (s.get('content', s.get('body', s.get('snippet', ''))) or '')[:300]
        else:
            title = getattr(s, 'title', '')[:80] if hasattr(s, 'title') else ''
            content = getattr(s, 'content', getattr(s, 'body', ''))[:300] if hasattr(s, 'content') else str(s)[:300]
        if content:
            line = f"{i+1}. {content}"
            if title:
                line = f"{i+1}. **{title}**: {content}"
            formatted.append(line)
    return "\n\n".join(formatted) if formatted else "未找到相关信息。"

# ----- 安全数学求值 -----
_ALLOWED_OP = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.Pow: operator.pow,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}

def _safe_eval(expr: str) -> int | float:
    tree = ast.parse(expr.strip(), mode='eval')
    def _eval(node):
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return node.value
            raise ValueError(f"不支持的常量: {type(node.value).__name__}")
        if isinstance(node, ast.BinOp):
            op = _ALLOWED_OP.get(type(node.op))
            if op is None:
                raise ValueError(f"不支持的操作符: {type(node.op).__name__}")
            return op(_eval(node.left), _eval(node.right))
        if isinstance(node, ast.UnaryOp):
            op = _ALLOWED_OP.get(type(node.op))
            if op is None:
                raise ValueError(f"不支持的操作符: {type(node.op).__name__}")
            return op(_eval(node.operand))
        if isinstance(node, ast.Expression):
            return _eval(node.body)
        raise ValueError(f"不支持的语法: {type(node).__name__}")
    return _eval(tree.body)

# ----- 工具2：计算器 -----
@tool
def calculator(expression: str) -> str:
    """数学计算。"""
    try:
        if not re.match(r'^[\d\+\-\*\/\.\(\)\s]+$', expression):
            return "非法字符"
        result = _safe_eval(expression)
        return f"{expression} = {result}"
    except Exception as e:
        return f"计算出错：{str(e)}"

# ----- 工具3：笔记系统 -----
NOTES_FILE = "notes.txt"

@tool
def save_note(content: str) -> str:
    """保存笔记。"""
    try:
        timestamp = dt.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(NOTES_FILE, "a", encoding="utf-8") as f:
            f.write(f"\n--- {timestamp} ---\n{content}\n")
        logger.info(f"笔记已保存: {content[:50]}...")
        return "✅ 笔记已保存"
    except Exception as e:
        return f"保存失败：{str(e)}"

@tool
def read_notes() -> str:
    """读取所有笔记。"""
    try:
        if not os.path.exists(NOTES_FILE):
            return "暂无笔记。"
        with open(NOTES_FILE, "r", encoding="utf-8") as f:
            content = f.read()
        return content[:5000] if len(content) > 5000 else content
    except Exception as e:
        return f"读取失败：{str(e)}"

@tool
def search_notes(keyword: str) -> str:
    """搜索笔记中的关键词。"""
    try:
        if not os.path.exists(NOTES_FILE):
            return "暂无笔记。"
        with open(NOTES_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
        matches = [line.strip() for line in lines if keyword in line]
        if not matches:
            return f"未找到包含「{keyword}」的笔记。"
        return "\n".join(matches[:20])
    except Exception as e:
        return f"搜索失败：{str(e)}"

# ----- 工具4：知识库 -----
@tool
def query_knowledge_base(query: str) -> str:
    """查询本地知识库。仅搜索已预先导入的本地文档数据，不包含实时新闻和互联网信息。如需查询新闻或最新动态请使用 web_search 工具。"""
    vectorstore = load_knowledge_base()
    if not vectorstore:
        return "知识库为空。"
    docs = vectorstore.similarity_search(query, k=3)
    return "\n\n".join([f"📄 {d.page_content[:300]}..." for d in docs]) if docs else "未找到相关信息。"

# ----- 工具5：Excel -----
@tool
def excel_tool(action: str, filename: str, data: str = "") -> str:
    """操作 Excel。"""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return "请安装 openpyxl"
    ensure_output_dir()
    filepath = os.path.join(OUTPUT_DIR, filename if filename.endswith('.xlsx') else filename + '.xlsx')
    if action == 'create':
        wb = Workbook()
        ws = wb.active
        if data:
            for row in data.strip().split('\n'):
                ws.append([c.strip() for c in row.split(',')])
        wb.save(filepath)
        return f"✅ Excel 已保存到 {filepath}"
    elif action == 'read':
        if not os.path.exists(filepath): return "文件不存在"
        wb = load_workbook(filepath)
        ws = wb.active
        return "\n".join([", ".join([str(c) if c else "" for c in row]) for row in ws.iter_rows(values_only=True)])
    elif action == 'add':
        if not os.path.exists(filepath): return "文件不存在"
        wb = load_workbook(filepath)
        ws = wb.active
        if data:
            for row in data.strip().split('\n'):
                ws.append([c.strip() for c in row.split(',')])
            wb.save(filepath)
            return f"✅ 已追加到 {filepath}"
        return "未提供数据"
    return "不支持的 action"

# ----- 工具6：读文件 -----
@tool
def read_file(filepath: str) -> str:
    """读取文件内容。返回文件全文或前 30000 字符（以较小者为准）。"""
    try:
        safe_path = _validate_path(filepath)
        if not os.path.exists(safe_path): return "文件不存在"
        with open(safe_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content[:30000] if len(content) > 30000 else content
    except PermissionError as e:
        return str(e)
    except Exception as e:
        return f"读取失败：{str(e)}"

# ----- 工具7：文件夹管理 -----
@tool
def manage_folder(action: str, path: str) -> str:
    """管理文件夹。"""
    try:
        safe_path = _validate_path(path)
        if action == 'list':
            return "\n".join(os.listdir(safe_path)) if os.path.exists(safe_path) else "路径不存在"
        elif action == 'create':
            os.makedirs(safe_path, exist_ok=True)
            return f"✅ 已创建 {safe_path}"
        elif action == 'delete':
            os.rmdir(safe_path) if os.path.exists(safe_path) else None
            return f"✅ 已删除 {safe_path}"
    except PermissionError as e:
        return str(e)
    except Exception as e:
        return f"操作失败：{str(e)}"

# ----- 工具8：系统信息 -----
@tool
def system_info(request: str = "all") -> str:
    """系统信息。"""
    try:
        cpu = f"CPU: {psutil.cpu_percent(interval=1)}%"
        mem = psutil.virtual_memory()
        mem_s = f"内存: {mem.used//(1024**2)}MB/{mem.total//(1024**2)}MB ({mem.percent}%)"
        disk = psutil.disk_usage('/')
        disk_s = f"磁盘: {disk.used//(1024**3)}GB/{disk.total//(1024**3)}GB ({disk.percent}%)"
        if request == 'cpu': return cpu
        if request == 'memory': return mem_s
        if request == 'disk': return disk_s
        return f"{cpu}\n{mem_s}\n{disk_s}"
    except Exception as e:
        return f"获取失败：{str(e)}"

# ----- 工具9：提醒（持久化）-----
REMINDER_FILE = "reminders.json"

def _load_reminders() -> list[dict]:
    if os.path.exists(REMINDER_FILE):
        try:
            with open(REMINDER_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _save_reminder_list(reminders: list[dict]) -> None:
    with open(REMINDER_FILE, "w", encoding="utf-8") as f:
        json.dump(reminders, f, ensure_ascii=False, indent=2)

def _check_pending_reminders() -> None:
    reminders = _load_reminders()
    now = time.time()
    for r in reminders:
        remaining = r["trigger_at"] - now
        if remaining > 0:
            def _fire(msg=r["message"], delay=remaining):
                time.sleep(delay)
                if os.name == 'nt':
                    import ctypes
                    ctypes.windll.user32.MessageBoxW(0, msg, "小想提醒", 0x40)
            threading.Thread(target=_fire, daemon=True).start()
            logger.info(f"恢复提醒：{r['message']}（{int(remaining)}秒后）")

@tool
def set_reminder(seconds: int, message: str = "时间到！") -> str:
    """设置提醒。"""
    trigger_at = time.time() + seconds
    def remind():
        time.sleep(seconds)
        if os.name == 'nt':
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, message, "小想提醒", 0x40)
    reminders = _load_reminders()
    # 清理已过期的提醒
    now = time.time()
    reminders = [r for r in reminders if r.get("trigger_at", 0) > now]
    reminders.append({"trigger_at": trigger_at, "message": message})
    _save_reminder_list(reminders)
    threading.Thread(target=remind, daemon=True).start()
    logger.info(f"提醒已设置：{seconds}秒后 - {message}")
    return f"✅ 已设置 {seconds} 秒后提醒"

@tool
def list_reminders() -> str:
    """查看所有待触发的提醒。"""
    reminders = _load_reminders()
    now = time.time()
    active = [r for r in reminders if r["trigger_at"] > now]
    if not active:
        return "暂无待触发的提醒。"
    lines = []
    for i, r in enumerate(active, 1):
        remaining = int(r["trigger_at"] - now)
        lines.append(f"{i}. 「{r['message']}」- 还剩 {remaining} 秒")
    return "\n".join(lines)

# ----- 工具10：剪贴板 -----
@tool
def clipboard(action: str, text: str = "") -> str:
    """读写剪贴板。"""
    if action == 'copy':
        if not text: return "请提供文本"
        pyperclip.copy(text)
        return "✅ 已复制"
    elif action == 'paste':
        content = pyperclip.paste()
        return content if content else "剪贴板为空"
    return "不支持的操作"

# ----- 工具11：图片下载 -----
@tool
def download_image(query: str, save_path: str = "") -> str:
    """根据关键词下载一张图片，保存在 output 文件夹。"""
    try:
        ensure_output_dir()
        if save_path:
            save_path = os.path.basename(save_path)
        if not save_path:
            safe = "".join(c for c in query if c.isalnum() or c in (' ', '_')).rstrip()
            save_path = f"{safe.replace(' ', '_')}.jpg"
        filepath = os.path.join(OUTPUT_DIR, save_path)
        try:
            from duckduckgo_search import DDGS
            with DDGS() as ddgs:
                results = list(ddgs.images(query, max_results=1))
            if results:
                image_url = results[0]['image']
                resp = requests.get(image_url, stream=True, timeout=15)
                if resp.status_code == 200:
                    with open(filepath, 'wb') as f:
                        for chunk in resp.iter_content(1024):
                            f.write(chunk)
                    return f"✅ 图片已保存到 {filepath}"
        except Exception:
            pass
        fallback_url = f"https://loremflickr.com/800/600/{query}"
        resp = requests.get(fallback_url, stream=True, timeout=15)
        if resp.status_code == 200:
            with open(filepath, 'wb') as f:
                for chunk in resp.iter_content(1024):
                    f.write(chunk)
            return f"✅ 图片已保存到 {filepath}（备用源）"
        return "下载失败：所有图片源均无响应"
    except Exception as e:
        return f"下载失败：{str(e)}"

# ----- 工具12：天气 -----
@tool
def get_weather(city: str) -> str:
    """查询指定城市的天气。city 为城市名称拼音或英文名（如 Beijing, Guangzhou）。"""
    try:
        city_map = {
            "北京": "Beijing", "上海": "Shanghai", "广州": "Guangzhou", "深圳": "Shenzhen",
            "杭州": "Hangzhou", "成都": "Chengdu", "武汉": "Wuhan", "南京": "Nanjing",
            "重庆": "Chongqing", "西安": "Xi'an", "长沙": "Changsha", "苏州": "Suzhou",
            "天津": "Tianjin", "厦门": "Xiamen", "青岛": "Qingdao", "大连": "Dalian"
        }
        city_en = city_map.get(city, city)
        geo_url = f"https://geocoding-api.open-meteo.com/v1/search?name={city_en}&count=1&language=zh"
        geo_resp = requests.get(geo_url, timeout=5)
        geo_data = geo_resp.json()
        if "results" not in geo_data or len(geo_data["results"]) == 0:
            return f"未找到城市「{city}」的天气信息"
        lat = geo_data["results"][0]["latitude"]
        lon = geo_data["results"][0]["longitude"]
        name = geo_data["results"][0].get("name", city)
        weather_url = f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}&current_weather=true"
        weather_resp = requests.get(weather_url, timeout=5)
        weather_data = weather_resp.json()
        current = weather_data["current_weather"]
        temp = current["temperature"]
        wind = current["windspeed"]
        desc = current.get("weathercode", 0)
        weather_desc = {
            0: "晴天", 1: "晴", 2: "多云", 3: "阴天",
            45: "有雾", 48: "霜雾", 51: "小毛毛雨", 53: "毛毛雨", 55: "中毛毛雨",
            61: "小雨", 63: "中雨", 65: "大雨",
            71: "小雪", 73: "中雪", 75: "大雪",
            80: "小阵雨", 81: "中阵雨", 82: "大阵雨",
            95: "雷阵雨", 96: "冰雹雷", 99: "强冰雹雷"
        }.get(desc, f"天气代码{desc}")
        return f"{name}当前天气：{weather_desc}，气温{temp}°C，风速{wind}km/h"
    except Exception as e:
        return f"查询天气失败：{str(e)}"

# ----- 工具13：翻译 -----
@tool
def translate_text(text: str, target: str = "zh") -> str:
    """翻译文本。"""
    try:
        return GoogleTranslator(source='auto', target=target).translate(text)
    except Exception as e:
        return f"翻译失败：{str(e)}"

# ----- 工具14：二维码 -----
@tool
def generate_qrcode(data: str, save_path: str = "qrcode.png") -> str:
    """生成二维码。"""
    try:
        ensure_output_dir()
        filepath = os.path.join(OUTPUT_DIR, save_path)
        qrcode.make(data).save(filepath)
        return f"✅ 二维码已保存到 {filepath}"
    except Exception as e:
        return f"生成失败：{str(e)}"

# ----- 工具15：写文件 -----
@tool
def write_file(filepath: str, content: str) -> str:
    """将内容写入文件。"""
    try:
        safe_path = _validate_path(filepath)
        os.makedirs(os.path.dirname(safe_path) or '.', exist_ok=True)
        with open(safe_path, 'w', encoding='utf-8') as f:
            f.write(content)
        return f"✅ 文件已保存：{safe_path}"
    except PermissionError as e:
        return str(e)
    except Exception as e:
        return f"写入失败：{str(e)}"

# ----- 工具16：安全替换函数（AST 语法树 + 自动备份）-----
@tool
def replace_function(filepath: str, function_name: str, new_function_code: str) -> str:
    """
    安全地替换指定文件中的某个函数。使用 AST 语法树精准定位。
    修改前会自动备份原文件到 backups 文件夹。
    """
    try:
        safe_path = _validate_path(filepath)
        backup_path = auto_backup(safe_path)
        if backup_path:
            logger.info(f"替换前已备份: {backup_path}")
        if not os.path.exists(safe_path):
            return f"❌ 文件不存在：{safe_path}"
        if not safe_path.endswith('.py'):
            return "❌ 只允许修改 .py 文件"
        with open(safe_path, 'r', encoding='utf-8') as f:
            source_lines = f.readlines()
        source_code = "".join(source_lines)
        clean_name = function_name.strip()
        if clean_name.startswith("def "):
            clean_name = clean_name[4:].strip()
        if '(' in clean_name:
            clean_name = clean_name.split('(')[0].strip()
        if '.' in clean_name:
            clean_name = clean_name.split('.')[-1].strip()
        tree = ast.parse(source_code)
        target_node = None
        if '.' in clean_name:
            # 支持 "ClassName.funcname" 格式，精准定位类方法
            class_name, method_name = clean_name.split('.', 1)
            for node in ast.walk(tree):
                if isinstance(node, ast.ClassDef) and node.name == class_name:
                    for item in node.body:
                        if isinstance(item, ast.FunctionDef) and item.name == method_name:
                            target_node = item
                            break
                    break
            if not target_node:
                return f"❌ 解析失败：未在类 '{class_name}' 中找到方法 '{method_name}'"
        else:
            for node in ast.walk(tree):
                if isinstance(node, ast.FunctionDef) and node.name == clean_name:
                    target_node = node
                    break
        if not target_node:
            return f"❌ 解析失败：未在文件中找到名为 '{clean_name}' 的函数"
        start_line = target_node.lineno - 1
        if hasattr(target_node, 'decorator_list') and target_node.decorator_list:
            start_line = target_node.decorator_list[0].lineno - 1
        end_line = target_node.end_lineno
        original_indent_str = source_lines[start_line][:len(source_lines[start_line]) - len(source_lines[start_line].lstrip())]
        dedented_new_code = textwrap.dedent(new_function_code.strip('\n'))
        indented_new_code = textwrap.indent(dedented_new_code, original_indent_str)
        final_new_lines = [line + '\n' for line in indented_new_code.split('\n')]
        new_source_lines = source_lines[:start_line] + final_new_lines + source_lines[end_line:]
        with open(safe_path, 'w', encoding='utf-8') as f:
            f.writelines(new_source_lines)
        return f"✅ 已通过 AST 精准替换函数 '{clean_name}'，文件 {safe_path} 已更新（备份已保存）"
    except SyntaxError as se:
        return f"❌ 替换失败：新代码存在语法错误 ({se})"
    except Exception as e:
        return f"❌ 替换报错：{str(e)}"

# ----- 工具17：待办列表 -----
TODOS_FILE = "todos.json"

def _load_todos() -> list[dict]:
    if os.path.exists(TODOS_FILE):
        try:
            with open(TODOS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def _save_todos(todos: list[dict]) -> None:
    with open(TODOS_FILE, "w", encoding="utf-8") as f:
        json.dump(todos, f, ensure_ascii=False, indent=2)

@tool
def add_todo(content: str) -> str:
    """添加一条待办事项。"""
    todos = _load_todos()
    todos.append({"id": len(todos) + 1, "content": content, "done": False})
    _save_todos(todos)
    logger.info(f"待办已添加: {content}")
    return f"✅ 已添加待办：{content}"

@tool
def list_todos() -> str:
    """查看所有待办事项。"""
    todos = _load_todos()
    if not todos:
        return "待办列表为空。"
    lines = []
    for t in todos:
        status = "✅" if t["done"] else "⬜"
        lines.append(f"{status} {t['id']}. {t['content']}")
    return "\n".join(lines)

@tool
def complete_todo(todo_id: str) -> str:
    """将待办标记为已完成。todo_id 为编号。"""
    todos = _load_todos()
    try:
        tid = int(todo_id.strip())
    except ValueError:
        return "❌ 请输入有效的编号"
    for t in todos:
        if t["id"] == tid:
            t["done"] = True
            _save_todos(todos)
            return f"✅ 已完成：{t['content']}"
    return f"❌ 未找到编号 {tid} 的待办"

@tool
def delete_todo(todo_id: str) -> str:
    """删除一条待办事项。todo_id 为编号。"""
    todos = _load_todos()
    try:
        tid = int(todo_id.strip())
    except ValueError:
        return "❌ 请输入有效的编号"
    for i, t in enumerate(todos):
        if t["id"] == tid:
            removed = todos.pop(i)
            _save_todos(todos)
            return f"✅ 已删除：{removed['content']}"
    return f"❌ 未找到编号 {tid} 的待办"

# ====== 插件系统 ======
PLUGINS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "plugins")

_DANGEROUS_MODULES = {"os", "subprocess", "shutil", "ctypes", "sys", "importlib", "socket", "requests"}

def _validate_plugin_safe(filepath: str) -> bool:
    """用 AST 检查插件文件是否包含危险 import"""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            tree = ast.parse(f.read())
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                for alias in node.names:
                    top_module = alias.name.split(".")[0]
                    if top_module in _DANGEROUS_MODULES:
                        logger.warning(f"插件 {filepath} 使用了危险模块：{alias.name}，已跳过")
                        return False
            elif isinstance(node, ast.ImportFrom):
                top_module = node.module.split(".")[0] if node.module else ""
                if top_module in _DANGEROUS_MODULES:
                    logger.warning(f"插件 {filepath} 使用了危险模块：{node.module}，已跳过")
                    return False
        return True
    except Exception as e:
        logger.error(f"插件安全检查失败 {filepath}: {e}")
        return False

PLUGIN_REGISTRY: list[tuple] = []

def _load_plugins() -> list:
    plugin_tools = []
    global PLUGIN_REGISTRY
    PLUGIN_REGISTRY = []
    if not os.path.exists(PLUGINS_DIR):
        os.makedirs(PLUGINS_DIR, exist_ok=True)
        return plugin_tools
    for fname in sorted(os.listdir(PLUGINS_DIR)):
        if not fname.endswith(".py") or fname.startswith("_"):
            continue
        fpath = os.path.join(PLUGINS_DIR, fname)
        if not _validate_plugin_safe(fpath):
            continue
        try:
            import importlib.util as _iu
            spec = _iu.spec_from_file_location(f"plugin_{fname[:-3]}", fpath)
            if not spec or not spec.loader:
                continue
            mod = _iu.module_from_spec(spec)
            spec.loader.exec_module(mod)
            if not hasattr(mod, "tool_meta") or not hasattr(mod, "handler"):
                logger.warning(f"插件 {fname} 缺少 tool_meta 或 handler，已跳过")
                continue
            meta = mod.tool_meta
            p_name = meta.get("name", f"plugin_{fname[:-3]}")
            p_desc = meta.get("description", "")
            p_cmd = meta.get("指令", "")
            p_has_param = meta.get("has_param", True)
            def _make_handler(m):
                def h(param: str = "") -> str:
                    try:
                        return str(m.handler(param))
                    except Exception as e:
                        return f"插件执行失败：{str(e)}"
                return h
            plugin_tool = Tool(
                name=p_name,
                description=p_desc,
                func=_make_handler(mod),
            )
            plugin_tools.append(plugin_tool)
            if p_cmd:
                PLUGIN_REGISTRY.append((p_cmd, p_name, p_desc, p_has_param))
            logger.info(f"✅ 插件已加载：{p_name}（{fname}）")
        except Exception as e:
            logger.error(f"插件加载失败 {fname}: {traceback.format_exc()}")
    return plugin_tools

# ====== Reasonix 代码智能体专用工具子集 ======
CODE_TOOLS = [
    read_file,
    write_file,
    replace_function,
]

# 启动时恢复未触发的提醒
_check_pending_reminders()

# 加载插件
_plugin_tools = _load_plugins()

# ====== 统一智能体调度入口 ======
@tool
def dispatch_agent(agent_name: str, task: str, mode: str = "normal") -> str:
    """将任务派给专家智能体处理。目前支持：Reasonix（代码专家，负责读/写/改/调试代码）。
       参数：
           agent_name: 智能体名称，目前仅支持 "Reasonix"
           task: 任务描述
           mode: 执行模式 — "normal"（按风险分级确认）或 "autonomous"（全自动执行，无需确认）
    """
    if agent_name not in ("Reasonix", "reasonix"):
        return f"❌ 未知智能体：{agent_name}。目前支持的智能体：Reasonix（代码专家）"
    try:
        from reasonix_agent import ReasonixAgent
        agent = ReasonixAgent()
        result = agent.run(task, mode=mode)
        return result
    except Exception as e:
        logger.error(f"Reasonix 调度异常: {traceback.format_exc()}")
        return f"❌ 智能体执行失败：{str(e)}"

@tool
def export_chat() -> str:
    """导出当前对话记录为 Markdown 文件。从 memory.json 读取并保存到 output/ 文件夹。"""
    import json as _json
    MEMORY_FILE = "memory.json"
    if not os.path.exists(MEMORY_FILE):
        return "❌ 暂无对话记录可导出"
    try:
        with open(MEMORY_FILE, "r", encoding="utf-8") as f:
            data = _json.load(f)
        if not data:
            return "❌ 对话记录为空"
        lines = ["# 与小想的对话\n"]
        for msg in data:
            role = msg.get("role", "unknown")
            content = msg.get("content", "")
            if role == "human":
                lines.append(f"**九里香**：{content}\n")
            elif role == "ai" or role == "assistant":
                lines.append(f"**小想**：{content}\n")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = "output"
        os.makedirs(out_dir, exist_ok=True)
        filepath = os.path.join(out_dir, f"对话_{ts}.md")
        with open(filepath, "w", encoding="utf-8") as f:
            f.writelines(lines)
        return f"✅ 对话已导出到 `{filepath}`"
    except Exception as e:
        return f"❌ 导出失败：{str(e)}"

# ====== 新增工具：护肤品成分分析 ======
@tool
def analyze_skincare(product_name: str, analysis_type: str = "safety") -> str:
    """分析护肤品的成分安全性和配伍禁忌。product_name 为产品名称，analysis_type 可选 'safety'(安全性) 或 'conflict'(与另一产品的冲突检查)。"""
    try:
        # Step 1: 搜索产品成分表
        search_query = f"{product_name} 全成分表 备案"
        if TAVILY_API_KEY:
            from langchain_tavily import TavilySearch
            search = TavilySearch(tavily_api_key=TAVILY_API_KEY, max_results=3, search_depth="advanced", include_answer=True)
            raw = search.invoke(search_query)
            if isinstance(raw, dict):
                sources = raw.get('results', [])
            else:
                sources = getattr(raw, 'results', [])
        else:
            from duckduckgo_search import DDGS
            with DDGS() as ddgs:
                raw = list(ddgs.text(search_query, max_results=3))
            sources = [{'title': r.get('title', ''), 'content': r.get('body', ''), 'url': r.get('href', '')} for r in raw]

        if not sources:
            return f"❌ 未找到「{product_name}」的成分信息。请确认产品名称是否正确，或尝试搜索其英文名。"

        # Step 2: 提取成分列表和来源
        ingredient_text = ""
        source_url = ""
        for s in sources:
            content = s.get('content', '') if isinstance(s, dict) else getattr(s, 'content', '')
            if '成分' in content or '备案' in content:
                ingredient_text = content[:1000]
                source_url = s.get('url', '') if isinstance(s, dict) else getattr(s, 'url', '')
                break
        if not ingredient_text:
            ingredient_text = sources[0].get('content', '')[:1000] if isinstance(sources[0], dict) else sources[0].content[:1000]
            source_url = sources[0].get('url', '') if isinstance(sources[0], dict) else getattr(sources[0], 'url', '')

        # Step 3: 根据分析类型调用 LLM 进行深度分析
        analysis_prompt = f"""你是一位资深化妆品配方师。请根据以下信息分析产品「{product_name}」：

成分信息来源：{source_url}
成分信息片段：{ingredient_text}

分析要求（请严格遵循）：
1. 列出该产品的主要功效成分及其作用。
2. 根据公开发表的成分安全性数据，标记出以下风险（如有）：
   - 致痘风险成分（如：肉豆蔻酸异丙酯、月桂醇聚醚-4、棕榈酸异丙酯等）
   - 刺激性成分（如：酒精/乙醇、香精、薄荷醇、高浓度酸类等）
   - 孕妇慎用成分（如：维A酸、视黄醇、水杨酸等）
3. 分析配方骨架：成分表前5位是否有较多硅油/增稠剂（可能为概念性添加）？
4. 如果产品采用包裹/缓释技术，请说明并据此下调风险等级。
5. 结合真实皮肤环境（pH缓冲、使用习惯）给出实际使用建议，避免纯理论化学反应推断。
6. 在分析末尾附上免责声明：\"以上分析基于公开成分数据和配方科学常识，仅供参考，不构成专业医疗建议。具体效果因人而异，建议先做局部测试。\"

请用自然段落输出，不要使用表格或 Markdown 格式。"""

        # 调用小想的 LLM 进行分析
        from langchain_core.messages import HumanMessage, SystemMessage
        from agent import llm
        analysis_result = llm.invoke([SystemMessage(content=analysis_prompt), HumanMessage(content="请开始分析。")])
        final_analysis = analysis_result.content.strip()

        # Step 4: 组合最终输出
        output = f"📱 **{product_name}** 成分分析\n\n"
        output += final_analysis
        output += f"\n\n🔗 成分来源：{source_url}"
        return output

    except Exception as e:
        logger.error(f"护肤品分析失败: {traceback.format_exc()}")
        return f"❌ 分析失败：{str(e)}。请检查产品名称或稍后重试。"

ALL_TOOLS = [
    web_search, calculator,
    save_note, read_notes, search_notes,
    query_knowledge_base, excel_tool,
    read_file, write_file, manage_folder,
    system_info,
    set_reminder, list_reminders,
    clipboard,
    download_image, get_weather, translate_text, generate_qrcode,
    add_todo, list_todos, complete_todo, delete_todo,
    replace_function,
    dispatch_agent,
    export_chat,
    analyze_skincare,  # 新增
] + _plugin_tools